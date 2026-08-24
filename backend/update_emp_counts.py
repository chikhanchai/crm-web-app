import openpyxl
import json

data = [
  {"company_name": "บริษัท อิปซอสส์ จำกัด", "employee_count_est": "200+"},
  {"company_name": "บริษัท อีไอบิซ จำกัด", "employee_count_est": "-"},
  {"company_name": "บริษัท เค-เน็กซ์ คอร์ปอเรชั่น จำกัด", "employee_count_est": "100-299"},
  {"company_name": "บริษัท ไวท์ไลน์แอคทิเวชัน จำกัด", "employee_count_est": "300+"},
  {"company_name": "บริษัท อชัวเรียน (ประเทศไทย) จำกัด", "employee_count_est": "-"},
  {"company_name": "บริษัท โนมิมาโช จำกัด", "employee_count_est": "-"},
  {"company_name": "บริษัท แอคทู-ลั่ม จำกัด", "employee_count_est": "100+"},
  {"company_name": "บริษัท จีวอฎัน (ประเทศไทย) จำกัด", "employee_count_est": "-"},
  {"company_name": "บริษัท เลิศวิลัยแอนด์ซันส์ จำกัด", "employee_count_est": "200"},
  {"company_name": "บริษัท อินะบาตะ ไทย จำกัด", "employee_count_est": "118"},
  {"company_name": "บริษัท ฟู้ดแพชชั่น จำกัด", "employee_count_est": "4000+"},
  {"company_name": "บริษัท เครือสุมิพล จำกัด", "employee_count_est": "50+"},
  {"company_name": "บริษัท เอส.เอ.ปิโตรเทค จำกัด", "employee_count_est": "100+"},
  {"company_name": "บริษัท ค้าเกษตรผล จำกัด", "employee_count_est": "-"},
  {"company_name": "บริษัท ไพน์ - แปซิฟิคคอร์ปอเรชั่น จำกัด", "employee_count_est": "50-99"},
  {"company_name": "บริษัท ออนวัลล่า จำกัด", "employee_count_est": "436"},
  {"company_name": "บริษัท อินเตอร์เนชั่นแนล ไร้ซ์ แอนด์ โปรดักซ์ จำกัด", "employee_count_est": "200+"},
  {"company_name": "บริษัท แอร์โค จำกัด", "employee_count_est": "500"},
  {"company_name": "บริษัท ฟู้ดโปรเจ็ค (สยาม) จำกัด", "employee_count_est": "400"},
  {"company_name": "บริษัท โซเนพาร์ (ประเทศไทย) จำกัด", "employee_count_est": "150"}
]

excel_path = r"C:\Happy\AI SandBox\CRM\Customer_Profiling_ICT_Solutions_Master_724Accounts_Strict_Web_Crosscheck2.xlsx"
wb = openpyxl.load_workbook(excel_path)
ws = wb.worksheets[0]

# Find the employee count column index from row 4
emp_col_idx = None
for col in range(1, ws.max_column + 1):
    header_val = ws.cell(row=4, column=col).value
    if header_val and ('พนักงาน' in str(header_val) or 'Est' in str(header_val)):
        emp_col_idx = col
        break

if not emp_col_idx:
    print("Could not find employee count column!")
    exit(1)

# Create lookup dictionary
emp_map = {item['company_name']: item['employee_count_est'] for item in data}

# Update Excel
updated = 0
for row in range(5, ws.max_row + 1):
    comp_name = ws.cell(row=row, column=4).value
    if comp_name and str(comp_name).strip() in emp_map:
        val = emp_map[str(comp_name).strip()]
        if val != '-':
            ws.cell(row=row, column=emp_col_idx).value = val
            updated += 1

wb.save(excel_path)
print(f"Successfully updated {updated} records.")
